#first install openpyxl using pip in the cmd!
#import the openpyxl library and get load workbook function to open an existing Excel .xlsx file.

import openpyxl
book = openpyxl.load_workbook(r"C:\Users\madhumitha\OneDrive\Desktop\projects\seleniumProjects\uploads_and_downloads\python_demo.xlsx")

#select the worksheet (sheet 1, sheet2,.... in the bottom)
sheet = book.active  #selects active sheet
#or you can get what sheet you want by using, book['sheet_name']

#next go to the cell you want
cell = sheet.cell(row=1, column=1).value

#to write in the cell
sheet.cell(row=2,column=2).value = "madhu"
sheet.cell(row=2,column=3).value = "mitha"

#to print the written cell
print(sheet.cell(row=2,column=3))
#or you can print by cell name
print(sheet['B2'].value)

#to get the maximum rows or maximum columns
total_rows, total_columns = sheet.max_row, sheet.max_column

#To print all the values from the sheet
#using for loop
for i in range(1, total_rows+1):
    for j in range(1, total_columns+1):
        print(sheet.cell(row=i, column=j).value)

#to print only the row which is "testcase2"
for i in range(1, total_rows+1):
    if sheet.cell(row=i,column=1).value == "testcase2":
        for j in range(1, total_columns+1):
            print(sheet.cell(row=i, column=j).value)

#print all this in dictionary
Dict = {}
for i in range(1, total_rows+1):
    if sheet.cell(row=i,column=1).value == "testcase2": #row =
        for j in range(2, total_columns+1):  #get the columns
            # Dict["f_name"] = "Madhumitha
             Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

print(Dict)

#save if you want to save something you wrote in Excel
book.save(r"C:\Users\madhumitha\OneDrive\Desktop\projects\seleniumProjects\uploads_and_downloads\python_demo.xlsx")











